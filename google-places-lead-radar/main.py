import os
import sys
import traceback
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# Google Places Lead-Radar
# -----------------------------------------------------------------------------
# Dieses Skript liest Suchbegriffe aus input.txt, fragt die offizielle Google
# Places API ab und speichert die Ergebnisse in einer formatierten Excel-Datei.
# -----------------------------------------------------------------------------

def get_project_dir() -> Path:
    """
    Ermittelt den Arbeitsordner.

    Normaler Python-Start:
        Ordner, in dem main.py liegt.

    Ausführbare PyInstaller-Datei:
        Ordner, in dem die .exe bzw. das ausführbare Programm liegt.
        Dadurch können .env, input.txt und output direkt neben der .exe liegen.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


PROJECT_DIR = get_project_dir()
REPO_ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = PROJECT_DIR / ".env" if getattr(sys, "frozen", False) else REPO_ROOT / ".env"
INPUT_FILE = PROJECT_DIR / "input.txt"
OUTPUT_DIR = PROJECT_DIR / "output"

API_URL = "https://places.googleapis.com/v1/places:searchText"

FIELD_MASK = (
    "places.id,"
    "places.displayName,"
    "places.formattedAddress,"
    "places.nationalPhoneNumber,"
    "places.websiteUri,"
    "places.primaryType,"
    "places.rating,"
    "places.userRatingCount,"
    "places.googleMapsUri,"
    "places.businessStatus"
)

COLUMN_ORDER = [
    "Suchbegriff",
    "Name",
    "Kategorie",
    "Adresse",
    "Telefonnummer",
    "Website",
    "Google Maps Link",
    "Bewertung",
    "Anzahl Bewertungen",
    "Geschäftsstatus",
    "Place ID",
    "Gefunden am",
]


def load_api_key() -> str:
    """Lädt den Google Places API-Key aus der .env-Datei."""
    if not ENV_FILE.exists():
        raise FileNotFoundError(
            f"Die Datei .env wurde nicht gefunden: {ENV_FILE}\n"
            "Kopieren Sie .env.example im Repository-Root, nennen Sie die Kopie .env "
            "und tragen Sie Ihren API-Key ein."
        )

    load_dotenv(ENV_FILE)
    api_key = os.getenv("API_KEY_GOOGLE_PLACES", "").strip()

    if not api_key:
        raise ValueError(
            "API_KEY_GOOGLE_PLACES fehlt oder ist leer. "
            "Bitte tragen Sie den API-Key in der Datei .env ein."
        )

    return api_key


def load_search_queries() -> list[str]:
    """Lädt Suchbegriffe aus input.txt. Kommentare und leere Zeilen werden ignoriert."""
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            "Die Datei input.txt wurde nicht gefunden. "
            "Bitte legen Sie input.txt im Projektordner an."
        )

    queries = []
    with INPUT_FILE.open("r", encoding="utf-8") as file:
        for line in file:
            cleaned_line = line.strip()
            if not cleaned_line:
                continue
            if cleaned_line.startswith("#"):
                continue
            queries.append(cleaned_line)

    if not queries:
        raise ValueError(
            "In input.txt wurden keine Suchbegriffe gefunden. "
            "Bitte tragen Sie mindestens einen Suchbegriff ein."
        )

    return queries


def search_google_places(api_key: str, query: str) -> list[dict]:
    """Sucht mit der Google Places API nach Unternehmen/Orten passend zur Suchanfrage."""
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": FIELD_MASK,
    }

    payload = {
        "textQuery": query,
        "languageCode": "de",
        "regionCode": "DE",
    }

    response = requests.post(API_URL, headers=headers, json=payload, timeout=30)

    if response.status_code != 200:
        raise RuntimeError(
            f"Google API Fehler bei Suchbegriff '{query}'. "
            f"Status-Code: {response.status_code}. Antwort: {response.text}"
        )

    data = response.json()
    return data.get("places", [])


def normalize_place(query: str, place: dict, found_at: str) -> dict:
    """Wandelt einen Google-Places-Treffer in eine Tabellenzeile um."""
    return {
        "Suchbegriff": query,
        "Name": place.get("displayName", {}).get("text", ""),
        "Kategorie": place.get("primaryType", ""),
        "Adresse": place.get("formattedAddress", ""),
        "Telefonnummer": place.get("nationalPhoneNumber", ""),
        "Website": place.get("websiteUri", ""),
        "Google Maps Link": place.get("googleMapsUri", ""),
        "Bewertung": place.get("rating", ""),
        "Anzahl Bewertungen": place.get("userRatingCount", ""),
        "Geschäftsstatus": place.get("businessStatus", ""),
        "Place ID": place.get("id", ""),
        "Gefunden am": found_at,
    }


def format_excel_file(output_file: Path) -> None:
    """Macht die Excel-Datei lesbarer: Filter, fixierte Kopfzeile, Tabelle und Spaltenbreiten."""
    workbook = load_workbook(output_file)
    worksheet = workbook.active
    worksheet.title = "Google Places Leads"

    max_row = worksheet.max_row
    max_column = worksheet.max_column

    # Erste Zeile fixieren: Beim Scrollen bleiben die Spaltenüberschriften sichtbar.
    worksheet.freeze_panes = "A2"

    # Autofilter für die komplette Ergebnistabelle.
    worksheet.auto_filter.ref = worksheet.dimensions

    # Kopfzeile optisch hervorheben.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Excel-Tabelle mit Filter-Dropdowns und lesbarer Formatierung erstellen.
    table_ref = f"A1:{get_column_letter(max_column)}{max_row}"
    table = Table(displayName="GooglePlacesLeads", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)

    # Spaltenbreiten sinnvoll setzen.
    preferred_widths = {
        "Suchbegriff": 28,
        "Name": 34,
        "Kategorie": 22,
        "Adresse": 44,
        "Telefonnummer": 20,
        "Website": 38,
        "Google Maps Link": 38,
        "Bewertung": 12,
        "Anzahl Bewertungen": 18,
        "Geschäftsstatus": 18,
        "Place ID": 30,
        "Gefunden am": 20,
    }

    for column_index in range(1, max_column + 1):
        header = worksheet.cell(row=1, column=column_index).value
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = preferred_widths.get(header, 18)

        for row_index in range(2, max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Links klickbar machen, wenn URLs vorhanden sind.
    link_columns = ["Website", "Google Maps Link"]
    header_to_column = {
        worksheet.cell(row=1, column=column_index).value: column_index
        for column_index in range(1, max_column + 1)
    }

    for header in link_columns:
        column_index = header_to_column.get(header)
        if not column_index:
            continue
        for row_index in range(2, max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    # Zahlenformate für Bewertungen und Bewertungsanzahl.
    rating_column = header_to_column.get("Bewertung")
    review_count_column = header_to_column.get("Anzahl Bewertungen")

    if rating_column:
        for row_index in range(2, max_row + 1):
            worksheet.cell(row=row_index, column=rating_column).number_format = "0.0"

    if review_count_column:
        for row_index in range(2, max_row + 1):
            worksheet.cell(row=row_index, column=review_count_column).number_format = "0"

    # Kopfzeile etwas höher, damit sie visuell sauber wirkt.
    worksheet.row_dimensions[1].height = 22

    workbook.save(output_file)


def save_to_excel(rows: list[dict]) -> Path:
    """Speichert die gesammelten Daten als formatierte Excel-Datei im output-Ordner."""
    OUTPUT_DIR.mkdir(exist_ok=True)

    timestamp_for_filename = datetime.now().strftime("%Y-%m-%d_%H%M")
    output_file = OUTPUT_DIR / f"google_places_leads_{timestamp_for_filename}.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(COLUMN_ORDER)
    for row in rows:
        worksheet.append([row.get(column, "") for column in COLUMN_ORDER])
    workbook.save(output_file)
    format_excel_file(output_file)

    return output_file


def main() -> None:
    print("Google Places Lead-Radar startet...")

    try:
        api_key = load_api_key()
        queries = load_search_queries()
    except Exception as error:
        print(f"Fehler: {error}")
        return

    all_rows = []
    found_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"{len(queries)} Suchbegriffe gefunden.")

    for index, query in enumerate(queries, start=1):
        print(f"[{index}/{len(queries)}] Suche: {query}")

        try:
            places = search_google_places(api_key, query)
        except Exception as error:
            print(f"Warnung: Suchbegriff wurde übersprungen. Grund: {error}")
            continue

        for place in places:
            all_rows.append(normalize_place(query, place, found_at))

        print(f"  {len(places)} Treffer gefunden.")

    if not all_rows:
        print("Es wurden keine Ergebnisse gefunden. Es wurde keine Excel-Datei erstellt.")
        return

    output_file = save_to_excel(all_rows)

    print("Fertig.")
    print(f"{len(all_rows)} Ergebnisse gespeichert.")
    print(f"Excel-Datei: {output_file}")


def wait_before_exit() -> None:
    """Hält das Konsolenfenster offen, wenn das Programm als .exe gestartet wurde."""
    if getattr(sys, "frozen", False):
        input("\nDrücken Sie Enter zum Beenden...")


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"\nUnerwarteter Fehler: {error}")
        traceback.print_exc()
    finally:
        wait_before_exit()
